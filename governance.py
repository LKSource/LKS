from datetime import datetime

import pandas as pd
from dateutil.utils import today

from govconfig import segments
from openpyxl import load_workbook
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from DPSDataGen import dpsdatagen
import win32com.client

test = True
notifications = {'PBSvsAlcon': True, 'Demand': True, 'Milestones': False}

def send_email(segment, body):
    """

    :param subject:
    :param table_style_html:
    :param message:
    :param body:
    """
    if not notifications[segment]: return
    with open('table_style.txt') as file:
        table_style_html = file.read()
    email_message = "<p style='"'color:#003366; font-family: Calibri; font-size:14.5px; text-align:left;'"'><br>Msg<br><br></p>"
    footer = "<p style='"'color:#003366;font-family: Calibri; font-size:12px;text-align:left;'"'><br><br>--------------------------------Auto Generated Message --------------------------------<br><br></p>"
    ol = win32com.client.Dispatch("outlook.application")
    olmail_item = 0x0  #size of the new email
    new_email = ol.CreateItem(olmail_item)
    new_email.Subject = segments[segment]['subject']
    if test:
        new_email.To = 'xxx@xxx.com'
    else:
        new_email.To = segments[segment]['to']
        new_email.CC = segments[segment]['cc']

    #newmail.Body = 'Hello, this is a test email.'
    email_message = email_message.replace("Msg", segments[segment]['message'])
    final_html = table_style_html + email_message + body + footer
    new_email.HTMLBody = final_html

    # attach='C:\\Users\\admin\\Desktop\\Python\\Sample.xlsx'
    # new_email.Attachments.Add(attach)

    # To display the mail before sending it
    # new_email.Display()

    new_email.Send()


def getData(segment):
    """

    :param segment:
    :return:
    """
    df = pd.read_excel('Output.xlsx', segment)
    #df = df[(df["PBS-ALCON Effort"] >= 0.5) | (df["PBS-ALCON Effort"] <= -0.5)]
    table_html = df.to_html(table_id='gov_table', index=False)
    table_html = table_html.replace('class="dataframe" id="gov_table"', 'class="styled-table" id="gov_table"')
    print(table_html)
    return table_html

def genOutput():
    path = 'Output.xlsx'
    df_dict = {}
    #print(segments.keys())
    for key in segments.keys():
        #end_email(key, getData(key))

        df = pd.read_excel('Input.xlsx', key)

        # Perform Data Manipulation
        if key == 'Demand':
            #df = df[(df["Reference - Fulfilment Status"] is None) & (df["Onsite/Offshore"] == 'Offshore')]
            df = df[['Account','Project Title','# of Demands','Primary Skill/s','Secondary Skill/s','Exp Level/JL','Location','Demands']][(df["Onsite/Offshore"] == 'Offshore') & (df["Reference - Fulfilment Status"].isnull())]
            segments[key]['message'] = segments[key]['message'].replace('Cnt', str(df.shape[0]))
        elif key == 'PBSvsAlcon':
            df = df[(df["PBS-ALCON Effort"] >= 0.5) | (df["PBS-ALCON Effort"] <= -0.5)]
        elif key == 'Milestones':
            options = ['name', 'name1']
            df['Initial Scheduled Date'] = pd.to_datetime(df['Initial Scheduled Date'])
            #df['Initial Scheduled Date'] = pd.to_datetime(df['Initial Scheduled Date'],format="DD-MM-YYYY")
            df = df[(df["DM MailID"].isin(options)) & (df["Milestone Status"] != "Posted to SAP") & (df['Initial Scheduled Date'] < datetime.today())][['Master Project Code','Milestone No','Initial Scheduled Date','Milestone Status','Milestone Amount in USD']].groupby(['Master Project Code','Milestone No','Initial Scheduled Date','Milestone Status']).sum().reset_index()
            print(df)
            #df = df[['Master Project Code','Initial Scheduled Date','Milestone Status','Milestone Amount in USD']][['Master Project Code','Initial Scheduled Date','Milestone Status','Milestone Amount in USD']].groupby(['Master Project Code','Initial Scheduled Date','Milestone Status']).sum()
            #df = df.groupby(['Master Project Code','Initial Scheduled Date','Milestone Status','Milestone Amount in USD']).sum()
        df_dict[key] = df

    #print(df_dict)
    with pd.ExcelWriter(path) as writer:
        for key in segments.keys():
            df_dict[key].to_excel(writer, sheet_name=key, index=False)
            #writer = pd.ExcelWriter('Output.xlsx')
        #writer.book = book

        # Write Data to Excel File
        #book = load_workbook('Output.xlsx')
'''
path = 'pandas_to_excel.xlsx'

with pd.ExcelWriter(path) as writer:
    writer.book = openpyxl.load_workbook(path)
    df.to_excel(writer, sheet_name='new_sheet1')
    df2.to_excel(writer, sheet_name='new_sheet2')
'''

def checkSegments():
    """

    """
    genOutput()
    #getData('PBSvsAlcon')
    for key in segments.keys():
        send_email(key, getData(key))


'''   
    print("inside check")
    #df = pd.read_csv('report.csv')
    df = pd.read_excel('Output.xlsx', 'PBSvsAlcon')
    table_html = df.to_html(table_id='gov_table', index=False)
    table_html = table_html.replace('class="dataframe" id="gov_table"', 'class="styled-table" id="gov_table"')
    print(table_html)
    with open('table_style.txt') as file:
        table_style_html = file.read()
    email_message = "Pls fix the Alcon vs PBS gap for following projects on priority"
    #final_html = table_style_html + email_message + table_html'''
'''with open('new_html_file.html','w') as html_file:
        html_file.write(final_html)
        print(final_html)'''
'''subject = "Comcast - PBs-Alcon volume"
    send_email(subject, table_style_html, email_message, table_html)
    #print(email_message)
'''
