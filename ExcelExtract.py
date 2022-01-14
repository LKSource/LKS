import openpyxl
import os.path
import re

from os import path
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles.borders import Border
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, Font, fills
from tempfile import gettempdir


def main():
    readColumn = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 27, 28, 29, 186, 305, 306, 344, 391, 392, 399, 442, 522, 535, 536, 552, 811, 928, 989, 1025, 1055, 1104, 1115]
    totalRows = 6000
    print("i am here") 
    filenum = 47
    rpt = Workbook()
    Report = rpt.create_sheet("Report", 0)

    iRows = 1

    defects = load_workbook('D:\\NBCU\\NBCU_47-P.xlsx', data_only=True)
    defectSheet = defects.worksheets[0]
        
    print("Inside the function")
    
    Report.cell(iRows, 1).value = "File Number"
    for i in range(0, len(readColumn)):        
        Report.cell(iRows, i + 2).value = defectSheet.cell(1, readColumn[i]).value
        print(i)
    defects.close()
    
    iRows += 1
    
    for fn in range(44, 48):
        defects = load_workbook('D:\\NBCU\\NBCU_' + str(fn) + '-P.xlsx', data_only=True)
        defectSheet = defects.worksheets[0]
        
        for j in range(2, totalRows):
            if defectSheet.cell(j, 1).value != None:
                Report.cell(iRows, 1).value = fn
                for i in range(0, len(readColumn)):        
                    Report.cell(iRows, i + 2).value = defectSheet.cell(j, readColumn[i]).value
                iRows += 1
        defects.close()
        print("Done : ", fn)
            
    print("Saved")
    rpt.save('D:\\NBCU\\Defects.xlsx')
    rpt.close() 
    

main()
