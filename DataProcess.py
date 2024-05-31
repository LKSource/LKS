import openpyxl
import calendar
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from _sqlite3 import Row
from openpyxl.chart import BarChart, Series, Reference, LineChart
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.chart.data_source import NumDataSource

font = Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
wb = openpyxl.load_workbook('TQMMetricData.xlsx',data_only=True)
wbr = openpyxl.Workbook()

report = wbr.create_sheet("Report", 0)
chart = wbr.create_sheet("Charts",1)

#print((wb.get_sheet_names()))
#sheet = wb.get_sheet_by_name('DataSheet')

iRow = 0
iColumn = 0
iRowSize = 0
iColumnSize = 1
iReportRow = 2
isExists = False
lTracks = ['Compass','Devices','SS','Apps','Platforms']
iData = 0
iAvg = 0
iDesCol = 0

lSource = [7,8,19,23,24,20,14,15,16,30,31,42,46,47,43,48,49,61,62,73,77,78,74,68,69,70,102,103]
lDestination = [2,3,4,5,6,7,8,9,10,12,13,14,15,16,17,18,19,21,22,23,24,25,26,27,28,29,31,32]
dIndex = {"Compass" : 0, "Devices" : 33, "SS" : 66, "Apps" : 99, "Platforms" : 132}

dReport = {0: [3,12], 1: [12,24], 2: [24,27], 3: [27,30], 4: [30,33]}

lMetricCol = [3,4,5,6,7]
lReportCol = ['Apr14-Dec14','Jan15-Dec15','Jan-Mar16', 'Apr-Jun16', 'Jul-Sep16']
lReportRow = ['QAintakes Count','Manual Test Case Preparation #','Total Test Execution #',
'Manual Test Execution productivity','Automation Test Execution productivity','Total Test Execution Productivity',
'Total Regression Test #','Test Case automatable #','Test case Automated #','QAintakes Count','Manual Test Case Preparation #',
'Total Test Execution #','Manual Test Execution productivity','Automation Test Execution productivity',
'Total Test Execution Productivity','Automation %','Automation Utilization %','QAintakes Count','Manual Test Case Preparation #',
'Total Test Execution #','Manual Test Execution productivity','Automation Test Execution productivity',
'Total Test Execution Productivity','Total Regression Test #','Test Case automatable #','Test case Automated #',
'Automation %','Automation Utilization %']

for colTitle in lReportCol:
    report.cell(row=1, column=lMetricCol[iDesCol]).value = colTitle
    iDesCol += 1

for iKey, iValue in dReport.items():
    print("Key %d" % iKey)
    for i in range(iValue[0], iValue[1]):
        print(i)

for iTrack in lTracks:
    print(iTrack)
    sheet = wb.get_sheet_by_name(iTrack)
    report.cell(row=lDestination[0] + dIndex[iTrack] -1, column=1).value = iTrack
    for iDestination in range (0, len(lDestination)):
        print(" Source : %d Destination : %d" % (lSource[iDestination], lDestination[iDestination] + dIndex[iTrack]))
        report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=2).value = lReportRow[iDestination]
        for iKey, iValue in dReport.items():
            iData = 0
            iAvg = 0
            for iCol in range(iValue[0], iValue[1]):
                if ((sheet.cell(row=lSource[iDestination], column=iCol).value != None) and (sheet.cell(row=lSource[iDestination], column=iCol).value != '#DIV/0!')):
                    iData = iData + float(sheet.cell(row=lSource[iDestination], column=iCol).value)
                    iAvg += 1
                #print(sheet.cell(row=lSource[iDestination] + dIndex[iTrack], column=iCol).value)
            avgData = 0
            if iData == 0:
                print(iData)
                avgData = round(iData,2)
            else:
                print(iData/iAvg)
                avgData = round(iData / iAvg,2)

            if '%' in lReportRow[iDestination]:
                report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=lMetricCol[iKey]).value = avgData * 100
            else:
                report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=lMetricCol[iKey]).value = avgData
#            print(" Source : %d Destination : %d" % (lSource(iDestination),lDestination(iDestination)))

lCharts = [0,9,3,12,6,15,17,20,23,26]
lChart = ['Test Case Volume (Closed)', 'Test Case Volume (Actual)', 'Productivity (Closed)', 'Productivity (Actual)',
          'Automation Volume', 'Automation', 'Test Case Volume', 'Productivity', 'Automation Volume', 'Automation']
minCol = 2;
maxCol = 7;

for iChart in range(len(lCharts)):
    if lChart[iChart] != 'Automation':
        chart1 = BarChart()
    else:
        chart1 = LineChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.y_axis.minorUnit = 10
    chart1.y_axis.majorUnit = 10
    chart1.y_axis.scaling.min = 1
 
    scalMax = 10
    
    if 'Volume' in lChart[iChart]:
        #chart1.y_axis.scaling.max = 100000.0
        chart1.y_axis.scaling.logBase = 10
    elif 'Productivity' in lChart[iChart]:
        chart1.y_axis.scaling.min = 0
        #chart1.y_axis.scaling.max = 100
        chart1.y_axis.minorUnit = 2
    #else:
        #chart1.y_axis.scaling.max = 100
    
    #txt = openpyxl.chart.text.Text("Test Case Volume (By Closed QA Intakes)")
    chart1.title = lChart[iChart]
    print(lChart[iChart])
    chart1.height = 6
    chart1.width = 12
    #chart1.plot_area.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0,h=0, w=0))
    
    #chart1.plot_area.barChart.

    print("iChart %s", (iChart))

    # setup and append the first series
    values = Reference(report, min_row = lDestination[lCharts[iChart]], min_col = minCol, max_row= lDestination[lCharts[iChart]], max_col=maxCol)
    labels = Reference(report, min_row = lDestination[lCharts[iChart]], min_col = minCol, max_row= lDestination[lCharts[iChart]], max_col=maxCol)
    series = Series(values, title_from_data=True)
    
    for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]], max_col=maxCol, max_row=lDestination[lCharts[iChart]]):
        if float(dat.value) > scalMax:
            scalMax = dat.value
    
    chart1.series.append(series)
    #chart1.series[0]. = DataLabelList((1,2,3,4))

    # setup and append the second series
    values = Reference(report, min_row = lDestination[lCharts[iChart]+1], min_col = minCol, max_row= lDestination[lCharts[iChart]+1], max_col=maxCol)
    series = Series(values, title_from_data=True)
    # series.labels = dlbl

    for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]+1], max_col=maxCol, max_row=lDestination[lCharts[iChart]+1]):
        if float(dat.value) > scalMax:
            scalMax = dat.value
    
    chart1.series.append(series)

    if lChart[iChart] != 'Automation':
        # setup and append the second series
        values = Reference(report, min_row = lDestination[lCharts[iChart]+2], min_col = minCol, max_row= lDestination[lCharts[iChart]+2], max_col= maxCol)
        series = Series(values, title_from_data=True)
        # series.labels = dlbl
        for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]+2], max_col= maxCol, max_row=lDestination[lCharts[iChart]+2]):
            if float(dat.value) > scalMax:
                scalMax = dat.value

        chart1.series.append(series)
        
        print("Scale Max : ",scalMax)
        chart1.y_axis.scaling.max   = scalMax     

    #data = Reference(report, min_col=3, min_row=1, max_row=4, max_col=7)
    cats = Reference(report, min_col=minCol+1, max_col=maxCol, min_row=1, max_row=1)
    #chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    #    = Font(name='Courrier', size=8)
    chart1.legend.position = "b"
    chart1.legend.overlay = False
    #chart1.title.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0,h=0.1, w=1))
    chart1.legend.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0.8,h=0.1, w=1))
    chart1.legend.graphicalProperties = GraphicalProperties()
    #chart1.legend.legendEntry[0].font = Font(size=24, italic=True)
    chart1.shape = 2

    if (iChart % 2) == 0:
        sCell = "A" + str((15 * (iChart // 2)+5))
        print(sCell)
        chart.add_chart(chart1, sCell)
    else:
        sCell = "H" + str((15 * (iChart // 2)+5))
        print(sCell)
        chart.add_chart(chart1, sCell)

wbr.save('Report.xlsx')


'''

for row in sheet.rows:
    iRowSize = iRowSize + 1
iRowSize = iRowSize + 1

while sheet.cell(row = 1, column = iColumnSize).value != None :
    report.cell(row = 1, column = iColumnSize).value = sheet.cell(row = 1, column = iColumnSize).value
    #report.cell(row = 2, column = iColumnSize).value = sheet.cell(row = 2, column = iColumnSize).value
    iColumnSize = iColumnSize + 1

report.cell(row = 1, column = iColumnSize).value = "Effort"

print(("Row Size %d", iRowSize))
print(("Column Size %d", iColumnSize))

for iRow in range(2, iRowSize):
    #isExists = False
    for i in range(2, iReportRow):
        #print(("iRow %d i %d", iRow, i))
        if sheet.cell(row = iRow, column = 1).value == report.cell(row = i, column = 1).value and sheet.cell(row = iRow, column = 2).value == report.cell(row = i, column = 2).value and sheet.cell(row = iRow, column = 3).value == report.cell(row = i, column = 3).value and sheet.cell(row = iRow, column = 6).value == report.cell(row = i, column = 6).value:
            report.cell(row = i, column = 11).value = sheet.cell(row = iRow, column = 11).value + report.cell(row = i, column = 11).value
            #print(("Matched"))
            #print(("Matched iRow %d i %d", iRow, i))
            isExists = True
            break
        else:
            isExists = False
        #    iReportRow = iReportRow + 1
    if isExists == False:
        #print(("Not Matched iReportRow %d iRow %d", iReportRow, iRow))
        for j in range(1, iColumnSize):
            report.cell(row = iReportRow, column = j).value = sheet.cell(row = iRow, column = j).value
        iReportRow = iReportRow + 1
    #else:
    #    report.cell(row = i, column = 11).value = sheet.cell(row = iRow, column = 11).value + report.cell(row = i, column = 11).value


for iRow in range(2, iReportRow):
    report.cell(row = iRow, column = iColumnSize).value = float(report.cell(row = iRow, column = 11).value) / float((calendar.monthrange(report.cell(row = iRow, column = 2).value, report.cell(row = iRow, column = 1).value)[1]))

isExists = False

summary = wbr.create_sheet("Summary", 0)

iSummaryRow = 3

summary.cell(row = 2, column = 1).value = "Master Code"
summary.cell(row = 2, column = 2).value = "Child Code"
for col in range(1,13):
    summary.cell(row = 1, column = ((col * 2) + 1)).value = calendar.month_name[col]
    summary.cell(row = 2, column = ((col * 2) + 1)).value = "Onsite"
    summary.cell(row = 2, column = ((col * 2) + 2)).value = "OffShore"


for iRow in range(2, iReportRow+1):
    #isExists = False
    for i in range(3, iSummaryRow):
        #print(("iRow %d i %d", iRow, i))
        if report.cell(row = iRow, column = 8).value == summary.cell(row = i, column = 1).value and report.cell(row = iRow, column = 7).value == summary.cell(row = i, column = 2).value:
            col = report.cell(row = iRow, column = 1).value
            location = report.cell(row = iRow, column = 6).value
            if location == "Onsite":
                effort = summary.cell(row = i, column = (col * 2) + 1).value
                if effort is None:
                    summary.cell(row = i, column = (col * 2) + 1).value = float(report.cell(row = iRow, column = 20).value)
                else:
                    summary.cell(row = i, column = (col * 2) + 1).value = float(report.cell(row = iRow, column = 20).value) + effort
            else:
                effort = summary.cell(row = i, column = (col * 2) + 2).value
                if effort is None:
                    summary.cell(row = i, column = (col * 2) + 2).value = float(report.cell(row = iRow, column = 20).value)
                else:
                    summary.cell(row = i, column = (col * 2) + 2).value = float(report.cell(row = iRow, column = 20).value) + effort
            #+ float(summary.cell(row = i, column = 20).value)
            #print(("Matched"))
            #print(("Matched iRow %d i %d", iRow, i))
            isExists = True
            break
        else:
            isExists = False
        #    iReportRow = iReportRow + 1
    if isExists == False:
        #print(("Not Matched iReportRow %d iRow %d", iReportRow, iRow))
        for j in range(1, iColumnSize):
            summary.cell(row = iSummaryRow, column = 1).value = report.cell(row = iRow, column = 8).value
            summary.cell(row = iSummaryRow, column = 2).value = report.cell(row = iRow, column = 7).value
            col = report.cell(row = iRow, column = 1).value
            location = report.cell(row = iRow, column = 6).value
            if location == "Onsite":
                effort = summary.cell(row = iSummaryRow, column = (col * 2) + 1).value
                if effort is None:
                    summary.cell(row = iSummaryRow, column = (col * 2) + 1).value = float(report.cell(row = iRow, column = 20).value)
            elif location == "OffShore":
                effort = summary.cell(row = iSummaryRow, column = (col * 2) + 2).value
                if effort is None:
                    summary.cell(row = iSummaryRow, column = (col * 2) + 2).value = float(report.cell(row = iRow, column = 20).value)
        iSummaryRow = iSummaryRow + 1

#wb.save('ALCON_Report.xlsx')
wbr.save('Report.xlsx')
print(("Process Completed ..."))
'''
