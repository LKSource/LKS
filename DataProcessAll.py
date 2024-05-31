import openpyxl
import calendar
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from _sqlite3 import Row
from openpyxl.chart import BarChart, Series, Reference, LineChart
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.chart.data_source import NumDataSource

font = Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
wb = openpyxl.load_workbook('TQMMetricData.xlsx',data_only=True)
wbr = openpyxl.Workbook()

report = wbr.create_sheet("Report", 0)
allChart = wbr.create_sheet("All", 1)
compassChart = wbr.create_sheet("Compass", 2)
devicesChart = wbr.create_sheet("Devices", 3)
ssChart = wbr.create_sheet("Streaming Services", 4)
etvAppsChart = wbr.create_sheet("eTV Apps", 5)
platformsChart = wbr.create_sheet("Platforms", 6)
#print((wb.get_sheet_names()))
#sheet = wb.get_sheet_by_name('DataSheet')

iRow = 0
iColumn = 0
iRowSize = 0
iColumnSize = 1
iReportRow = 2
isExists = False
lTracks = ['Compass','Devices','SS','Apps','Platforms','All']
iData = 0
iAvg = 0
iDesCol = 0

lSource = [7,8,19,23,24,20,14,15,16,30,31,42,46,47,43,48,49,61,62,73,77,78,74,68,69,70,102,103]
lDestination = [2,3,4,5,6,7,8,9,10,12,13,14,15,16,17,18,19,21,22,23,24,25,26,27,28,29,31,32]
dIndex = {"Compass" : 0, "Devices" : 33, "SS" : 66, "Apps" : 99, "Platforms" : 132, "All" : 165}

dReport = {0: [3,12], 1: [12,24], 2: [24,27], 3: [27,30], 4: [30,33], 5: [33,36], 6: [36,39], 7: [39,41]}

lMetricCol = [3,4,5,6,7,8,9,10]
lReportCol = ['Apr14-Dec14','Jan15-Dec15','Jan-Mar16', 'Apr-Jun16', 'Jul-Sep16', 'Oct-Dec16', 'Jan-Mar17', 'Apr-May17']
lReportRow = ['QAintakes Count','Manual Test Case Preparation #','Total Test Execution #',
'Manual Test Execution productivity','Automation Test Execution productivity','Total Test Execution Productivity',
'Total Regression Test #','Test Case automatable #','Test case Automated #','QAintakes Count','Manual Test Case Preparation #',
'Total Test Execution #','Manual Test Execution productivity','Automation Test Execution productivity',
'Total Test Execution Productivity','Automation %','Automation Utilization %','QAintakes Count','Manual Test Case Preparation #',
'Total Test Execution #','Manual Test Execution productivity','Automation Test Execution productivity',
'Total Test Execution Productivity','Total Regression Test #','Test Case automatable #','Test case Automated #',
'Automation %','Automation Utilization %']

for colTitle in lReportCol:
    report.cell(row=1, column=lMetricCol[iDesCol]).value = colTitle
    iDesCol += 1

for iKey, iValue in dReport.items():
    print("Key %d" % iKey)
    for i in range(iValue[0], iValue[1]):
        print(i)

for iTrack in lTracks:
    print(iTrack)
    report.cell(row=lDestination[0] + dIndex[iTrack] -1, column=1).value = iTrack
    
    if iTrack == "All": break;
    
    sheet = wb.get_sheet_by_name(iTrack)
    for iDestination in range (0, len(lDestination)):
        print(" Source : %d Destination : %d" % (lSource[iDestination], lDestination[iDestination] + dIndex[iTrack]))
        report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=2).value = lReportRow[iDestination]
        for iKey, iValue in dReport.items():
            iData = 0
            iAvg = 0
            for iCol in range(iValue[0], iValue[1]):
                if ((sheet.cell(row=lSource[iDestination], column=iCol).value != None) and (sheet.cell(row=lSource[iDestination], column=iCol).value != '#DIV/0!')):
                    iData = iData + float(sheet.cell(row=lSource[iDestination], column=iCol).value)
                    iAvg += 1
                #print(sheet.cell(row=lSource[iDestination] + dIndex[iTrack], column=iCol).value)
            avgData = 0
            if iData == 0:
                print(iData)
                avgData = round(iData,2)
            else:
                print(iData/iAvg)
                avgData = round(iData / iAvg,2)

            if '%' in lReportRow[iDestination]:
                report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=lMetricCol[iKey]).value = avgData * 100
            else:
                report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=lMetricCol[iKey]).value = avgData
#            print(" Source : %d Destination : %d" % (lSource(iDestination),lDestination(iDestination)))

iTrack = "All"

for iDestination in range (0, len(lDestination)):
    print(" Source : %d Destination : %d" % (lSource[iDestination], lDestination[iDestination] + dIndex[iTrack]))
    report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=2).value = lReportRow[iDestination]
    
    for iCol in range (0, len(lMetricCol)):
        iData = 0
        iAvg = 0
        for jTrack in lTracks:
            if jTrack != "All":
                if ((report.cell(row=lDestination[iDestination] + dIndex[jTrack], column=lMetricCol[iCol]).value != None) and (report.cell(row=lDestination[iDestination] + dIndex[jTrack], column=lMetricCol[iCol]).value != '#DIV/0!')):
                    iData = iData + float(report.cell(row=lDestination[iDestination] + dIndex[jTrack], column=lMetricCol[iCol]).value)
                    iAvg += 1
            #print(sheet.cell(row=lSource[iDestination] + dIndex[iTrack], column=iCol).value)
        avgData = 0
        if iData == 0:
            print(iData)
            avgData = round(iData,2)
        else:
            print(iData/iAvg)
            avgData = round(iData / iAvg,2)

        #if '%' in lReportRow[iDestination]:
        #    report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=lMetricCol[iCol]).value = avgData * 100
        #else:
        report.cell(row=lDestination[iDestination] + dIndex[iTrack], column=lMetricCol[iCol]).value = avgData
#            print(" Source : %d Destination : %d" % (lSource(iDestination),lDestination(iDestination)))

lCharts = [0,9,3,12,6,15,17,20,23,26]
lChart = ['Test Case Volume (Closed)', 'Test Case Volume (Actual)', 'Productivity (Closed)', 'Productivity (Actual)',
          'Automation Volume', 'Automation', 'Test Case Volume', 'Productivity', 'Automation Volume', 'Automation']
minCol = 2;
maxCol = 10;

for iTrack in lTracks:    
    if iTrack == "Compass":
        chart = compassChart;
    elif iTrack == "Devices":
        chart = devicesChart;     
    elif iTrack == "SS":
        chart = ssChart;
    elif iTrack == "Apps":
        chart = etvAppsChart;
    elif iTrack == "Platforms":
        chart = platformsChart;
    elif iTrack == "All":
        chart = allChart;
                    
    for iChart in range(len(lCharts)):
        if lChart[iChart] != 'Automation':
            chart1 = BarChart()
        else:
            chart1 = LineChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.y_axis.minorUnit = 10
        chart1.y_axis.majorUnit = 10
        chart1.y_axis.scaling.min = 1
     
        scalMax = 10
        
        if 'Volume' in lChart[iChart]:
            #chart1.y_axis.scaling.max = 100000.0
            chart1.y_axis.scaling.logBase = 10
        elif 'Productivity' in lChart[iChart]:
            chart1.y_axis.scaling.min = 0
            #chart1.y_axis.scaling.max = 100
            chart1.y_axis.minorUnit = 2
        #else:
            #chart1.y_axis.scaling.max = 100
        
        #txt = openpyxl.chart.text.Text("Test Case Volume (By Closed QA Intakes)")
        chart1.title = lChart[iChart]
        print(lChart[iChart])
        chart1.height = 6
        chart1.width = 14
        #chart1.plot_area.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0,h=0, w=0))
        
        #chart1.plot_area.barChart.
    
        #print("iChart %s", (iChart))
    
        # setup and append the first series
        values = Reference(report, min_row = lDestination[lCharts[iChart]] + dIndex[iTrack], min_col = minCol, max_row= lDestination[lCharts[iChart]] + dIndex[iTrack], max_col=maxCol)
        labels = Reference(report, min_row = lDestination[lCharts[iChart]] + dIndex[iTrack], min_col = minCol, max_row= lDestination[lCharts[iChart]] + dIndex[iTrack], max_col=maxCol)
        series = Series(values, title_from_data=True)
        
        for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]] + dIndex[iTrack], max_col=maxCol, max_row=lDestination[lCharts[iChart]] + dIndex[iTrack]):
            if dat.value != None:
                if float(dat.value) > scalMax:
                    scalMax = dat.value
        
        chart1.series.append(series)
        #chart1.series[0]. = DataLabelList((1,2,3,4))
    
        # setup and append the second series
        values = Reference(report, min_row = lDestination[lCharts[iChart]+1] + dIndex[iTrack], min_col = minCol, max_row= lDestination[lCharts[iChart]+1] + dIndex[iTrack], max_col=maxCol)
        series = Series(values, title_from_data=True)
        # series.labels = dlbl
    
        for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]+1] + dIndex[iTrack], max_col=maxCol, max_row=lDestination[lCharts[iChart]+1] + dIndex[iTrack]):
            if dat.value != None:
                if float(dat.value) > scalMax:
                    scalMax = dat.value
        
        chart1.series.append(series)
    
        if lChart[iChart] != 'Automation':
            # setup and append the second series
            values = Reference(report, min_row = lDestination[lCharts[iChart]+2] + dIndex[iTrack], min_col = minCol, max_row= lDestination[lCharts[iChart]+2] + dIndex[iTrack], max_col= maxCol)
            series = Series(values, title_from_data=True)
            # series.labels = dlbl
            for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]+2] + dIndex[iTrack], max_col= maxCol, max_row=lDestination[lCharts[iChart]+2] + dIndex[iTrack]):
                if dat.value != None:
                    if float(dat.value) > scalMax:
                        scalMax = dat.value
    
            chart1.series.append(series)
            
            #print("Scale Max : ",scalMax)
            chart1.y_axis.scaling.max   = scalMax     
    
        #data = Reference(report, min_col=3, min_row=1, max_row=4, max_col=7)
        cats = Reference(report, min_col=minCol+1, max_col=maxCol, min_row=1, max_row=1)
        #chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        #    = Font(name='Courrier', size=8)
        chart1.legend.position = "b"
        chart1.legend.overlay = False
        #chart1.title.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0,h=0.1, w=1))
        chart1.legend.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0.8,h=0.1, w=1))
        chart1.legend.graphicalProperties = GraphicalProperties()
        #chart1.legend.legendEntry[0].font = Font(size=24, italic=True)
        chart1.shape = 2
    
        if (iChart % 2) == 0:
            sCell = "A" + str((15 * (iChart // 2)+5))
            #print(sCell)
            chart.add_chart(chart1, sCell)
        else:
            sCell = "J" + str((15 * (iChart // 2)+5))
            #print(sCell)
            chart.add_chart(chart1, sCell)

# Track level graphs 

'''
lCharts = [0,9,3,12,6,15,17,20,23,26]
lChart = ['Test Case Volume (Closed)', 'Test Case Volume (Actual)', 'Productivity (Closed)', 'Productivity (Actual)',
          'Automation Volume', 'Automation', 'Test Case Volume', 'Productivity', 'Automation Volume', 'Automation']
minCol = 2;
maxCol = 7;

for iChart in range(len(lCharts)):
    if lChart[iChart] != 'Automation':
        chart1 = BarChart()
    else:
        chart1 = LineChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.y_axis.minorUnit = 10
    chart1.y_axis.majorUnit = 10
    chart1.y_axis.scaling.min = 1
 
    scalMax = 10
    
    if 'Volume' in lChart[iChart]:
        #chart1.y_axis.scaling.max = 100000.0
        chart1.y_axis.scaling.logBase = 10
    elif 'Productivity' in lChart[iChart]:
        chart1.y_axis.scaling.min = 0
        #chart1.y_axis.scaling.max = 100
        chart1.y_axis.minorUnit = 2
    #else:
        #chart1.y_axis.scaling.max = 100
    
    #txt = openpyxl.chart.text.Text("Test Case Volume (By Closed QA Intakes)")
    chart1.title = lChart[iChart]
    print(lChart[iChart])
    chart1.height = 6
    chart1.width = 12
    #chart1.plot_area.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0,h=0, w=0))
    
    #chart1.plot_area.barChart.

    print("iChart %s", (iChart))

    # setup and append the first series
    values = Reference(report, min_row = lDestination[lCharts[iChart]], min_col = minCol, max_row= lDestination[lCharts[iChart]], max_col=maxCol)
    labels = Reference(report, min_row = lDestination[lCharts[iChart]], min_col = minCol, max_row= lDestination[lCharts[iChart]], max_col=maxCol)
    series = Series(values, title_from_data=True)
    
    for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]], max_col=maxCol, max_row=lDestination[lCharts[iChart]]):
        if float(dat.value) > scalMax:
            scalMax = dat.value
    
    chart1.series.append(series)
    #chart1.series[0]. = DataLabelList((1,2,3,4))

    # setup and append the second series
    values = Reference(report, min_row = lDestination[lCharts[iChart]+1], min_col = minCol, max_row= lDestination[lCharts[iChart]+1], max_col=maxCol)
    series = Series(values, title_from_data=True)
    # series.labels = dlbl

    for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]+1], max_col=maxCol, max_row=lDestination[lCharts[iChart]+1]):
        if float(dat.value) > scalMax:
            scalMax = dat.value
    
    chart1.series.append(series)

    if lChart[iChart] != 'Automation':
        # setup and append the second series
        values = Reference(report, min_row = lDestination[lCharts[iChart]+2], min_col = minCol, max_row= lDestination[lCharts[iChart]+2], max_col= maxCol)
        series = Series(values, title_from_data=True)
        # series.labels = dlbl
        for (dat,) in report.iter_cols(min_col = minCol+1, min_row = lDestination[lCharts[iChart]+2], max_col= maxCol, max_row=lDestination[lCharts[iChart]+2]):
            if float(dat.value) > scalMax:
                scalMax = dat.value

        chart1.series.append(series)
        
        print("Scale Max : ",scalMax)
        chart1.y_axis.scaling.max   = scalMax     

    #data = Reference(report, min_col=3, min_row=1, max_row=4, max_col=7)
    cats = Reference(report, min_col=minCol+1, max_col=maxCol, min_row=1, max_row=1)
    #chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    #    = Font(name='Courrier', size=8)
    chart1.legend.position = "b"
    chart1.legend.overlay = False
    #chart1.title.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0,h=0.1, w=1))
    chart1.legend.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0.8,h=0.1, w=1))
    chart1.legend.graphicalProperties = GraphicalProperties()
    #chart1.legend.legendEntry[0].font = Font(size=24, italic=True)
    chart1.shape = 2

    if (iChart % 2) == 0:
        sCell = "A" + str((15 * (iChart // 2)+5))
        print(sCell)
        chart.add_chart(chart1, sCell)
    else:
        sCell = "H" + str((15 * (iChart // 2)+5))
        print(sCell)
        chart.add_chart(chart1, sCell)
'''

wbr.save('Report.xlsx')
