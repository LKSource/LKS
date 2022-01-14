import openpyxl
import os.path
import re

from os import path
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles.borders import Border
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, Font, fills
from tempfile import gettempdir


def main():
    readColumn = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 27, 28, 29, 186, 305, 306, 344,
                  391, 392, 399, 442, 522, 535, 536, 552, 811, 928, 989, 1025, 1055, 1104, 1115]
    totalRows = 6000
    print("i am here")
    filenum = 47
    rpt = Workbook()
    Report = rpt.create_sheet("Report", 0)

    iRows = 1
    filename = 'D:\\lingasamy_k\\OneDrive - Infosys Limited\\LINGAA\\Work\\Comcast\\Data Collection\\DU Trackers\\Projects\\2020\\NBCU\\NBCUniversal_OrgChart_Dec_Joe.xlsx'
    defects = load_workbook(filename, data_only=True)

    defectSheet = defects.worksheets[3]

    print("Inside the function")
    lst = []
    for iRows in range(2, totalRows):
        if defectSheet.cell(iRows, 1).value != None:
            lst = (defectSheet.cell(iRows, 1).value).split(",")
            print(lst[2])
            Report.cell(iRows, 1).value = lst[0]
            Report.cell(iRows, 2).value = (lst[1].split())[0]
            Report.cell(iRows, 3).value = lst[2]
            Report.cell(iRows, 4).value = lst[3]
#            for i in range(0, len(readColumn)):
#                Report.cell(iRows, i + 2).value = defectSheet.cell(j, readColumn[i]).value
#            iRows += 1
    defects.close()

    print("Saved")
    rpt.save('D:\\Defects.xlsx')
    rpt.close()

main()
