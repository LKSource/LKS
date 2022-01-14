#lex_auth_01269438594930278448

def find_pairs_of_numbers(num_list,n):
    #Remove pass and write your logic here
    count = 0
    for num1 in num_list:
        for num2 in num_list:
            if (num1 + num2) == n:                
                count = count + 1
    return int(count/2)

num_list=[1, 2, 7, 4, 5, 6, 0, 3]
n=6
print(find_pairs_of_numbers(num_list,n))

_______________________________

#lex_auth_0127382164803993601239

#This verification is based on string match.

def sum_of_numbers(list_of_num,filter_func=None):
    #Remove pass and write the logic here
    rsum = 0
    if filter_func == None:
        for num in list_of_num:
            rsum += num
    else:
        for num in filter_func(list_of_num):
            rsum += num
    return rsum
    
def even(data):
    #Remove pass and write the logic here
    data_list = []
    for num in data:
        if num % 2 == 0:
            data_list.append(num)
    return data_list

def odd(data):
    #Remove pass and write the logic here
    data_list = []
    for num in data:
        if num % 3 == 0:
            data_list.append(num)
    return data_list

sample_data = range(1,11)

print(sum_of_numbers(sample_data,None))

____________________________

#lex_auth_01269441810970214471

def check_double(number):
    #Remove pass and write your logic here
    
    if list(map(int,str(number))).sort() == list(map(int,str(number*2))).sort():
        return True
    else:
        return False

#Provide different values for number and test your program
print(check_double(125874))

______________________

#lex_auth_012693816779112448160

def calculate(distance,no_of_passengers):
    #Remove pass and write your logic here
    if (no_of_passengers * 80) - ((distance/10) * 70) > 0:
        return (no_of_passengers * 80) - ((distance/10) * 70)
    else:
        return -1



#Provide different values for distance, no_of_passenger and test your program
distance=20
no_of_passengers=50
print(calculate(distance,no_of_passengers))

___________________

#lex_auth_01269438947391897654

#Global variable
list_of_marks=(12,18,25,24,2,5,18,20,20,21)

def find_more_than_average():
    #Remove pass and write your logic here
    sums = 0
    for mark in list_of_marks:
        sums += mark
    sums = sums / len(list_of_marks)
    count = 0
    for mark in list_of_marks:
        if mark > sums:
            count += 1
    return (count / len(list_of_marks)) * 100

def sort_marks():
    #Remove pass and write your logic here
    mark_list = list(list_of_marks)
    mark_list.sort()
    return mark_list

def generate_frequency():
    #Remove pass and write your logic here
    freq_list = []
    mark_list = list(list_of_marks)
    for num in mark_list:
        freq_list.append(mark_list.count(num))
    return (freq_list)

print(find_more_than_average())
print(generate_frequency())
print(sort_marks())

______________________

#lex_auth_01269437504257228837
from builtins import IndexError

def find_average(mark_list):
    total=0
    try:
        for i in range(0, len(mark_list)+1):
            total+=mark_list[i]
        marks_avg=total/len(mark_list)
        return marks_avg
    except TypeError:
        print("TypeError")
    except ZeroDivisionError:
        print("ZeroDivisionError")
    except IndexError:
        print("IndexError")

try:
    m_list=[]        
    print("Average marks:", find_average(m_list))
except ValueError:
    print("ValueError")
    
_________________________

def human_tower(no_of_people):
    if (no_of_people == 1):
        return 1*(50)
    else:
        return no_of_people*(50) + human_tower(no_of_people - 2)

print(human_tower(5))

_____________________

#lex_auth_01269437527007232044

def human_pyramid(no_of_people):
    #remove pass and place the recursive code the you had written earlier for this function
    if(no_of_people==1):
        return 1*(50)
    else:
        return no_of_people*(50)+human_pyramid(no_of_people-2)

def find_maximum_people(max_weight):
    no_of_people=1
    #write your logic here. You may invoke recursive function human_pyramid() wherever applicable
    weight = 0
    while weight < max_weight:
        no_of_people += 2
        weight = human_pyramid(no_of_people+2)
        
    return no_of_people

#Provide different values for max_weight and test your program
max_people=find_maximum_people(1000)
print(max_people)

________________________

#lex_auth_01269442114344550475

def is_palindrome(word):
    #Remove pass and write your logic here   

    if len(word) < 2:
        return True

    #print (word[0].lower(),word[len(word)-1].lower())        
    if word[0].lower() == word[len(word)-1].lower():
        if len(word[1:-1]) > 1:
            return True and is_palindrome(word[1:-1])
        else:
            return True
    else:
        return False

words = ["Noon","Racecar","Reporter","Radar","Reader","Redder","Refer","Repaper"]
#Provide different values for word and test your program

for word in words:
    result=is_palindrome(word)
    if(result):
        print(word + " : Palindrome")
    else:
        print(word + " : Not a Palindrome")
       
 ___________

#lex_auth_01269442963365068878

def find_factors(num):
    #Accepts a number and returns the list of all the factors of a given number
    factors = []
    for i in range(2,(num+1)):
        if(num%i==0):
            factors.append(i)
    return factors

def is_prime(num, i):
    #Accepts the number num and num/2 --> i and returns True if the number is prime ,else returns False
    if(i==1) or (num==2) or (num==3):
        return True           
    elif(num%i==0):
        return False;
    else:
        return(is_prime(num,i-1))

def find_largest_prime_factor(list_of_factors):
    #Accepts the list of factors and returns the largest prime factor
    largest_prime = 1
    for num in list_of_factors:
        if is_prime(num, 3):
            #print(num)
            largest_prime = num
    return largest_prime

def find_f(num):
    #Accepts the number and returns the largest prime factor of the number
    #print(find_largest_prime_factor(find_factors(num)))
    return (find_largest_prime_factor(find_factors(num)))

def find_g(num):
    #Accepts the number and returns the sum of the largest prime factors of the 9 consecutive numbers starting from the given number
    sumv = 0
    for i in range(0,9):
        sumv += find_f(num + i)
    return sumv

#Note: Invoke function(s) from other function(s), wherever applicable.

print(find_g(11))

_______


#lex_auth_01269442760027340879

def find_factors(num):
    factors = []
    for i in range(1,num+1):
        if (num%i ==0):
            factors.append(i)
    return factors

def find_smallest_number(num):
    #start writing your code here
    i=int(1)
    while(True):
        x=find_factors(i)
        if(len(x)==num):
            print(x)
            break
        else:
            i+=int(1)
    return x[-1]
    #start writing your code here    

num=15
print("The number of divisors :",num)
result=find_smallest_number(num)
print("The smallest number having",num," divisors:",result)

__________

# lex_auth_0127136303145779201199


def check_for_ten(num1, num2):
    # start writing your code here
    if num1 == 10 or num2 == 10 or num1 + num2 == 10:
        return True
    return False

    
print(check_for_ten(10, 9))

_______________

# lex_auth_0127136253311385601197


def check_occurence(string):
    # start writing your code here
    if string.lower().count('mat') == string.lower().count('jet') and string.lower().count('mat') > 0:
        return True
    else:
        return False

        
string = "Jet on the Mat but mat is too long"
print(check_occurence(string))

________________________

# lex_auth_0127136291543285761194


def list_of_count(word_list):
    # start writing your code here
    count_list = [len(n) for n in word_list]
    return count_list


word_list = ["COme", "here"]
print(list_of_count(word_list))

__________________

# lex_auth_0127136213490565121191


def find_gcd(num1, num2):
    # start writing your code here
    numl1 = [num for num in range(2, num1 + 1) if num1 % num == 0]
    numl2 = [num for num in range(2, num2 + 1) if num2 % num == 0]
    rslt = 0
    for num in numl1:
        if num in numl2:
            rslt = num
    return rslt


num1 = 10000
num2 = 20
print(find_gcd(num1, num2))

___________

# lex_auth_0127136251125350401190


def divisible_by_sum(number):
    # start writing your code here
    if number % sum(list(map(int, str(number)))) == 0:
        return True
    else: 
        return False

    
number = 42
print(divisible_by_sum(number))

_______________

# lex_auth_0127136075580375041172


def check_22(num_list):
    # start writing your code here
    result = list()
    indices = [index for index, element in enumerate(num_list) if element == 2]
    result = [True for i in range(len(indices) - 1, 0, -1) if indices[i] - indices[i - 1] == 1]
    if len(result) > 0:
        return result[0]
    else:
        return False


print(check_22([1, 4, 5, 2, 2, 9]))

____________

# lex_auth_0127136011356405761166


def generate_sentences(subjects, verbs, objects):
    # start writing your code here
    sentence_list = []
    for sub in subjects:
        for ver in verbs:
            for obj in objects:
                sentence_list.append(sub + " " + ver + " " + obj)
    
    return sentence_list


subjects = ["I", "You"]
verbs = ["love", "play"]
objects = ["Hockey", "Football"]
print(generate_sentences(subjects, verbs, objects))

__________

# lex_auth_0127136021907046401165
import re


def find_upper_and_lower(sentence):
    # start writing your code here
    result_list = []
    result_list.append(len(re.findall('[A-Z]', sentence)))
    result_list.append(len(re.findall('[a-z]', sentence)))
    
    return result_list


sentence = "Come Here"
print(find_upper_and_lower(sentence))

____________________

# lex_auth_0127135945621340161163


def string_both_ends(input_string):
    # start writing your code here
    if len(input_string) < 2: return -1
    return input_string[:2] + input_string[-2:] 


input_string = "w3resource"
print(string_both_ends(input_string))

_____

# lex_auth_0127135904626688001159


def generate_dict(number):
    # start writing your code here
    new_dict = {n:n * n for n in range(1, number + 1)}
    return new_dict


number = 20
print(generate_dict(number))

________

# lex_auth_0127135929511936001157


def calculate_net_amount(trans_list):
    # start writing your code here
    wList = [int(n.split(':')[1]) for n in [n for n in trans_list if n.count('W') > 0]]
    dList = [int(n.split(':')[1]) for n in [n for n in trans_list if n.count('D') > 0]]
    net_amount = sum(dList) - sum(wList)
    
    return net_amount


trans_list = ["D:300", "D:200", "W:200", "D:100"]
print(calculate_net_amount(trans_list))

_______

# lex_auth_0127135857243832321154

from functools import reduce


def seed_no(number, ref_no):
    # start writing your code here
    numlst = list(map(int, str(number)))
    number = number * reduce((lambda x, y: x * y), list(map(int, str(number))))
    if number == ref_no: return True
    return False

     
number = 123
ref_no = 738
print(seed_no(number, ref_no))

_____


# lex_auth_0127135869481369601150


def list123(nums):
    # start writing your code here
        
    if 1 in nums and 2 in nums and 3 in nums:
        one1 = nums.index(1)
        two2 = nums.index(2)
        three3 = nums.index(3)
        
        if (three3 - two2 == 1) and (two2 - one1 == 1): return True
    return False


nums = [1, 2, 4, 4, 3]
print(list123(nums))

____

# lex_auth_0127135838317445121147
import re


def count_digits_letters(sentence):
    # start writing your code here
    result_list = list()
    result_list.append(len(re.findall("[a-zA-Z]", sentence))) 
    result_list.append(len(re.findall("[0-9]", sentence)))
    return result_list


sentence = "Infosys Mysore 570027"
print(count_digits_letters(sentence))

_________

# lex_auth_0127135811649044481146

def find_nine(nums):
    # Remove pass and write your code here
    for i in range(len(nums)):
        if i > 3: break
        if nums[i] == 9 : return True
    return False


nums = [1, 9, 4, 5, 6]
print(find_nine(nums))

______

# lex_auth_0127135787454873601143


def create_new_dictionary(prices):
    # start writing your code here
    new_dict = dict()
    for k in prices.keys():
        if prices[k] > 200:
            new_dict[k] = prices[k]
    
    return new_dict


prices = { 'ACME': 45.23, 'AAPL': 612.78, 'IBM': 205.55, 'HPQ': 37.20, 'FB': 10.75}
print(create_new_dictionary(prices))

_____

# lex_auth_0127135773590405121141

import re


def bracket_pattern(input_str):
    # Remove pass and write your code here
    res = list(filter(lambda x:  x in ('(', ')'), input_str))
    
    if res[0] == ')' or res[len(res) - 1] == '(': return False
    
    res1 = list(filter(lambda x:  x in '(', input_str))
    res2 = list(filter(lambda x:  x in ')', input_str))
    
    if len(res1) == len(res2) : return True
    
    return False

    
input_str = "(())("
print(bracket_pattern(input_str))

_______

# lex_auth_0127135739583692801137


def add_string(str1):
  # start writing your code here
  if len(str1) < 3 : return str1
  if str1.count("ing") == 0: 
      str1 += "ing"
  elif str1[-3:] == "ing":
      str1 += "ly"      
  
  return str1


str1 = "is"
print(add_string(str1))

______

# lex_auth_01269445968039936095


def validate_credit_card_number(card_number):
    # start writing your code here
    cnum = list(map(int, str(card_number)))
    cnum1 = [cnum[i] for i in range(len(cnum)) if i % 2 == 0]
    cnum2 = [cnum[i] for i in range(len(cnum)) if i % 2 != 0]
    cnum1 = [cnum1[i] * 2 for i in range(len(cnum1))] 
    # print([sum(list(map(int, str(cnum1[i])))) for i in range(len(cnum1)) if cnum1[i] > 9])
    
    for i in range(len(cnum1)):
        if cnum1[i] > 9: cnum1[i] = sum(list(map(int, str(cnum1[i]))))
    
    if (sum(cnum1) + sum(cnum2)) % 10 == 0: 
        return True
    else:
        return False


card_number = 1456734512345698  # 4539869650133101  #1456734512345698 # #5239512608615007
result = validate_credit_card_number(card_number)
if(result):
    print("credit card number is valid")
else:
    print("credit card number is invalid")

_____
    
# lex_auth_012694465248329728100

import re


def validate_name(name):
    # Start writing your code here
    # Name should not be empty, name should not exceed 15 characters
    # Name should contain only alphabets    
    if name.strip() == '': return False
    if len(name) > 15: return False
    if not name.isalpha(): return False
    return True

    
def validate_phone_no(phno):
    # Start writing your code here
    # Phone number should have 10 digits
    # Phone number should not have any characters or special characters
    # All the digits of the phone number should not be same.
    # Example: 9999999999 is not a valid phone number    
    try:
        if len(phno) != 10: return False
        if not phno.isdigit(): return False
#        num = int(phno) 
        if phno.count(phno[0]) == 10: return False    
        return True
    except ValueError:
        return False


def validate_email_id(email_id):
    # Start writing your code here
    '''It should contain one '@' character and '.com'
    '.com' should be present at the end of the email id.
    Domain name should be either 'gmail', 'yahoo' or 'hotmail'''
    regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
    domain = ["gmail", "yahoo", "hotmail"]
    
    if email_id.count('@') != 1: return False
    if email_id.count('.com') != 1: return False    
    if email_id[-4:] != '.com': return False
    if email_id[email_id.index('@') + 1:-4] not in domain: return False
    if not re.search(regex, email_id): return False
    return True


def validate_all(name, phone_no, email_id):
    # Start writing your code here
    # Use the below given print statements to display appropriate messages
    # Also, do not modify them for verification to work
    if not validate_name(name):
        print("Invalid Name")
    elif not validate_phone_no(phone_no):
        print("Invalid phone number")          
    elif not validate_email_id(email_id):
        print("Invalid email id")
    else:    
        print("All the details are valid")
    return


# Provide different values for name, phone_no and email_id and test your program
validate_all("Tina", "9999999908", "tina@gmail.com")

______

#lex_auth_01269446157664256093

def check_prime(number):
    #remove pass and write your logic here. if the number is prime return true, else return false
    if number == 1: return False
    if number == 2 or number == 3:
        return True
    else:
        for i in range (2,number):
            if number % i == 0 :
                return False
    return True

def rotations(num):
    #remove pass and write your logic here. It should return the list of different combinations of digits of the given number.
    #Rotation should be done in clockwise direction. For example, if the given number is 197, the list returned should be [197, 971, 719]
    numlist = []
    if len(str(num)) == 1:
        numlist.append(num)
        return numlist
    
    divv = 1
    for i in range(1,len(str(num))):
        divv *= 10
    #print(divv)        
    
    for i in range(0, len(str(num))):
        numlist.append(num)
        num = (num % divv) * 10 + num // divv
    #print(numlist)
    return numlist

def get_circular_prime_count(limit):
    #remove pass and write your logic here.It should return the count of circular prime numbers below the given limit.
    count = 0
    for i in range(1,limit+1):
        check = True
        for j in rotations(i):
            if not check_prime(j):
                check = False
        if check:
            #print(i)
            count += 1
    return count
#Provide different values for limit and test your program
print(get_circular_prime_count(100))
#print(rotations(100))
_____

#lex_auth_01269443664174284882

def ispalindrome(number):
    iput = number
    oput = 0
    
    while number !=0 :
    #    print(number%10, number//10)
        oput = oput * 10 + (number%10)
        number = int(number//10)
        
    #print(input, output)
    if iput == oput:
        return True
    else:
        return False

def nearest_palindrome(number):
    #start writitng your code here
    
    while True:
        number += 1
        if (ispalindrome(number)):
            return number

number=12331
print(nearest_palindrome(number))
_____

#lex_auth_0127382193364008961449

#Sample ticket list - ticket format: "flight_no:source:destination:ticket_no"
#Note: flight_no has the following format - "airline_name followed by three digit number

#Global variable
ticket_list=["AI567:MUM:LON:014","AI077:MUM:LON:056", "BA896:MUM:LON:067", "SI267:MUM:SIN:145","AI077:MUM:CAN:060","SI267:BLR:MUM:148","AI567:CHE:SIN:015","AI077:MUM:SIN:050","AI077:MUM:LON:051","SI267:MUM:SIN:146"]

def find_passengers_flight(airline_name="AI"):
    #This function finds and returns the number of passengers travelling in the specified airline.
    count=0
    for i in ticket_list:
        string_list=i.split(":")
        if(string_list[0].startswith(airline_name)):
            count+=1
    return count

def find_passengers_destination(destination):
    #Write the logic to find and return the number of passengers traveling to the specified destination
    count=0
    for i in ticket_list:
        string_list=i.split(":")
        if(string_list[2] == destination):
            count+=1
    return count

def find_passengers_per_flight():
    '''Write the logic to find and return a list having number of passengers traveling per flight based on the details in the ticket_list
    In the list, details should be provided in the format:
    [flight_no:no_of_passengers, flight_no:no_of_passengers, etc.].'''
    result = []
    for i in ticket_list:
        string_list=i.split(":")
        result.append(string_list[0] + ":" + string_list[3])

    return result    

def sort_passenger_list():  
    #Write the logic to sort the list returned from find_passengers_per_flight() function in the descending order of number of passengers
    result = find_passengers_per_flight()
    finalr = []
    passengerscnt = []
    for i in result:
        string_list=i.split(":")
        passengerscnt.append(string_list[1])
    passengerscnt.sort(key=int,reverse=True)
    
    for i in passengerscnt:
        for j in result:
#            print(int(i),int(j.split(":")[1]))
            if int(i) == int(j.split(":")[1]):
                #print(j)
                finalr.append(j)
    return finalr
    
#Provide different values for airline_name and destination and test your program.
print(find_passengers_flight("AI"))
print(find_passengers_destination("LON"))
print(sort_passenger_list())
_________

# lex_auth_01269446533799116898


def getfactors(number):
    factors = []
    for i in range(1, number):
        if (number % i == 0):
            factors.append(i)
            
    return factors


def check_perfect_number(number):
    # start writing your code here
    if number == 0 : return False
    if sum(getfactors(number)) == number:
        return True
    else:
        return False
    

def check_perfectno_from_list(no_list):
    # start writing your code here
    result = list()
    for num in no_list:
        if check_perfect_number(num):
            result.append(num)
    return result


# perfectno_list=check_perfectno_from_list([18,6,4,2,1,28,496,99,8128,98])        
perfectno_list = check_perfectno_from_list([87, 76, 567, 99, 0])
print(perfectno_list)

________

#lex_auth_01269446319507046499

def remove_duplicates(value):
    #start writing your code here
    resultstr = []
    for i in range(0,len(value)):
        if value[i] not in resultstr:
            resultstr.append(value[i])
    return ''.join(resultstr)        

print(remove_duplicates("11223445566666ababzzz@@@123#*#*"))

_______

#lex_auth_0127382206342184961397

def check_anagram(data1,data2):
    #start writing your code here
    if (len(data1) != len(data2)):
        return False
    
    for i in range(0,len(data1)):
        if data1[i].lower() == data2[i].lower():
            return False
    data1 = list(data1.lower())
    data2 = list(data2.lower())
    data1.sort()
    data2.sort()
    if ''.join(data1) == ''.join(data2):
        return True
    else:
        return False

print(check_anagram("About","table"))

______

try:
    flight_file=open("flight.txt","r")
    text=flight_file.read()
    print(text)
    flight_file.write(",Good Morning")
except:
    print("Error occurred")
finally:
    print("File is being closed")
    flight_file.close()
    if flight_file.closed:
        print("File is closed")
    else:
        print("File is open")
        
        ______
        
        
#lex_auth_01269443477535129681

def find_duplicates(list_of_numbers):
    #start writing your code here
    list_of_duplicates = []
    for i in range(0,len(list_of_numbers)):
        dcount = 0
        for j in range(0,len(list_of_numbers)):
            if list_of_numbers[i] == list_of_numbers[j]:
                dcount+=1
        if (dcount) > 1:
            if list_of_numbers[i] not in list_of_duplicates:
                list_of_duplicates.append(list_of_numbers[i])
    return list_of_duplicates

list_of_numbers=[1,2,2,3,3,3,4,4,4,4]
#list_of_numbers=[12,54,68,759,24,15,12,68,987,758,25,69]
list_of_duplicates=find_duplicates(list_of_numbers)
print(list_of_duplicates)

_______

#lex_auth_01269442545042227276

def find_ten_substring(num_str):
    #Remove pass and write your logic here
    result_list = []
    for i in range(0, len(num_str)+1):
        num = 0
        numstr = ""
        j = i
        while num < 10 and j < len(num_str):            
            num += int(num_str[j])
            numstr += num_str[j]
            j+=1
            #print (num, numstr)
            
        if num == 10:
            result_list.append(numstr)
    return result_list
        
num_str="3523014"
print("The number is:",num_str)
result_list=find_ten_substring(num_str)
print(result_list)

_______

# lex_auth_0127136518921830401222


def find_target_positions(input_string, target_word):
    # target_list = []
    # Start writing your code here
    instr = input_string.split()
    target_list = [i for i in range(len(instr)) if instr[i] == target_word]
    return target_list


def find_inverted_index(input_string):
    target_dict = {}
    # Start writing your code here
    instr = input_string.split()
    
    for word in instr:
        if word not in target_dict.keys():
            target_dict[word] = find_target_positions(input_string, word)
    
    return target_dict
    
    
input_string = "we dont need no education we dont need no thought control no we dont"
print(find_target_positions(input_string, "dont"))
result_dict = find_inverted_index(input_string)
print(result_dict)

_____

# lex_auth_0127136447430656001216

bmarks = {1000:"M", 100:"C", 10:"X", 1:"I", 500:"D", 50:"L", 5:"V"}
tens = [1000, 100, 10, 1]


def getnumeral(number, pos):
    rslt = ""
    num = number // pos
    if num == 4 and pos == 1000:
        for i in range(num): rslt += bmarks[pos]
    elif num in (4, 9):
        rslt += outbound(num, pos)
    elif num == 5:
        rslt += bmarks[pos * 5]
    elif num in range(1, 4):
        for i in range(num): rslt += bmarks[pos]
    elif num in range(6, 9):
        rslt += bmarks[pos * 5]
        for i in range(num - 5): rslt += bmarks[pos]
   
    return rslt


def outbound(num, pos):
    rslt = ""
    if num == 4:
        keys = list(bmarks.keys())
        rslt = bmarks[pos] + bmarks[pos * 5]
    elif num == 9:
        keys = list(bmarks.keys())        
        rslt = bmarks[keys[keys.index(pos)]] + bmarks[keys[keys.index(pos) - 1]]
    return rslt


def convert_to_roman(number):
    # Start writing your code here

    if number not in range(1, 5000): return -1    
    rslt = ""
    
    for i in tens:
        rslt += getnumeral(number, i)
        number = number % i
   
    return (rslt) 


# print(" : ", convert_to_roman(475))
for num in range(1, 10):    
    print(num, " : ", convert_to_roman(num))

    ______
    
# lex_auth_0127136426924195841210


def check_well_formatted(input_item, list_label):
    # Start writing your code here
    if type(input_item) != list : return False
    if len(input_item) < 2: return False
    if input_item[0] not in list_label: return False
    rslt = True
    for n in input_item[1:]:
        # print("n : ", n)        
        if type(n) == list: 
            rslt = check_well_formatted(n, list_label)
        elif (type(n) != str) : return False

    return rslt


input_list1 = ['NP', ['N', 'a', 'or', 'b'], 'c']
list_label1 = ['NP', 'V', 'N']
result = check_well_formatted(input_list1, list_label1)
if result is True:
    print("Well formatted")
else:
    print("Not Well formatted")

______

# lex_auth_0127136373866577921209


def integer_to_english(number):
    # start writing your code here
    ones = ["one", "two", "three", "four", "five", "six", "seven", "eight", "nine"]
    tens = ["ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen", "seventeen", "eighteen", "nineteen"]
    ties = ["", "twenty", "thirty", "fourty", "fifty", "sixty", "seventy", "eighty", "ninety"]
    more = ["hundred", "thousand"]
    if number not in range(1, 1001): return -1    
    rslt = ""
    while number:
        num = number // 1000
        if num > 0:
            rslt += ones[num - 1] + " " + more[1]
        number = number % 1000
        
        num = number // 100
        if num > 0:
            rslt += ones[num - 1] + " " + more[0]
        number = number % 100
        
        num = number 
        if num > 0:
            if rslt != "": rslt += " and"             
            if num in range (10, 20):
                if rslt != "": rslt += " " 
                rslt += tens[num - 10]
            else:
                num = number // 10
                if num > 0: 
                    if rslt != "": rslt += " "
                    rslt += ties[num - 1] 
                if (number % 10) > 0 : 
                    if rslt != "": rslt += " "
                    rslt += ones[(number % 10) - 1]
        number = 0
    return (rslt)
    

number = 1001
print(integer_to_english(number))

____


# lex_auth_0127136357122129921205
from math import sqrt


def check_squares(number):
    # start writing your code here
    for i in range(1, int(sqrt(number))):
         for j in range(1, int(sqrt(number))):
             if number == (i * i) + (j * j) : return True
    return False


number = 68
print(check_squares(number))

______

# lex_auth_0127136209566679041189


def check_numbers(num1, num2):
    # start writing your code here
    
    nlist = [n for n in range(num1, num2 + 1)]
    
    # print(nlist)
    
    num_list = []
    
    for i in range(num1, num2 + 1):
        for j in nlist:
            if i % j == 0 and i != j: 
                if i not in num_list : num_list.append(i)
                break
    num_list = set(num_list)
    return [num_list, len(num_list)]


num1 = 10
num2 = 30
print(check_numbers(num1, num2))

____

# lex_auth_0127136084107673601177


def rotate_list(input_list, n):
    # start writing your code here
    output_list = []
    lstsize = len(input_list)
    for i in range(0, len(input_list)):
        output_list.append(input_list[i - n])
        # print (input_list[i - n])
    return output_list


input_list = [1, 2, 3, 4, 5, 6]
output_list = rotate_list(input_list, 4)
print(output_list)

_____

# lex_auth_0127136008767324161169


def close_number(num1, num2, num3):
    # start writing your code here
    rslt = []
    rslt.append(abs(num1 - num2))
    rslt.append(abs(num2 - num3))
    rslt.append(abs(num3 - num1))
    
    if 1 in rslt : rslt.remove(1)
    elif 0 in rslt: rslt.remove(0)       
    else: return False
    
    cnt = 0
    for i in rslt: 
        if i == 2 or i > 2: cnt += 1
    if cnt == 2: return True
        
    return False

    
print(close_number(5, 6, 7))

_____

# lex_auth_0127136332814499841204


def sum_of_elements(num_list, number):
    ind = []
    for i in range(0, len(num_list)):
        
        if i < 1:
            if num_list[i] != number and num_list[i + 1] != number:
                ind.append(num_list[i])
        elif i == len(num_list) - 1:              
            if num_list[i] != number and num_list[i - 1] != number:
                ind.append(num_list[i])
        else:                            
            if num_list[i] != number and num_list[i + 1] != number and num_list[i - 1] != number:
                ind.append(num_list[i])

    return sum(ind)

      
num_list = [1, 7, 3, 4, 1, 7, 10, 5]
number = 7
print(sum_of_elements(num_list, number))

______

# lex_auth_0127136357873991681201


def exchange_list(number_list):
    # start writing your code here
    j = len(number_list) - 1 
    for i in range(0, len(number_list) // 2):
        temp = number_list[i] 
        number_list[i] = number_list[j]
        number_list[j] = temp
        # print(number_list)
        j -= 1
        
    # print (len(number_list) // 2, len(number_list))
    midnum = len(number_list) // 2        
    num = (midnum + ((len(number_list) - midnum) // 2))
    if len(number_list) % 2 != 0:
        midnum += 1    
    
    j = len(number_list) - 1 
    for i in range(midnum, num):
        temp = number_list[i] 
        number_list[i] = number_list[j]
        number_list[j] = temp
        # print(number_list)
        j -= 1    
    
    return number_list

     
number_list = [1, 2, 3, 4, 5, 6, 7, 8]
print(exchange_list(number_list))

_______

#pickletest

import pickle

def dmp():
    dogs_dict = { 'Ozzy': 3, 'Filou': 8, 'Luna': 5, 'Skippy': 10, 'Barco': 12, 'Balou': 9, 'Laika': 16 }
    filename = 'dogs'
    outfile = open(filename, 'wb')
    pickle.dump(dogs_dict,outfile)
    outfile.close()

def lod():
    dogs_dict = {}
    filename = 'dogs'
    infile = open(filename,'rb')
    dogs_dict = pickle.load(infile)
    print(dogs_dict)
    infile.close()

def main():
    dmp()
    lod()

if __name__ == "__main__":
    # calling main function
    main()

    _____

#PinCodePickup
import openpyxl
import os.path
import re

from os import path
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles.borders import Border
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, Font, fills
from tempfile import gettempdir


def getMaxRows(sheet):
    iRows = 1
    while sheet.cell(iRows, 1).value != None:
        iRows += 1
    print (iRows)
    return iRows    
'''    for iRows in range (1, sheet.max_row):
        # print(sheet.cell(iRows,1).value)
        if (sheet.cell(iRows, 1).value == None):
            return int(iRows)
'''


def main():
    addrss = load_workbook('Address.xlsx', data_only=True)
    rpt = Workbook()
    Report = rpt.create_sheet("Report", 0)
    AddSheet = addrss.worksheets[0]

    iRows = 1
    tRows = getMaxRows(AddSheet)
    print("tRows", tRows)
    # sheet.max_row
    
    Report.cell(iRows, 1).value = AddSheet.cell(1, 1).value
    Report.cell(iRows, 2).value = AddSheet.cell(1, 2).value
    Report.cell(iRows, 3).value = AddSheet.cell(1, 3).value
    Report.cell(iRows, 4).value = AddSheet.cell(1, 4).value
    Report.cell(iRows, 5).value = AddSheet.cell(1, 5).value
    Report.cell(iRows, 6).value = "New Pincode"
    
    for i in range(2, tRows):
        for j in range(1, 6):
            Report.cell(i, j).value = AddSheet.cell(i, j).value
        pincode = (AddSheet.cell(i, 4).value).split(',')
        # print(pincode)
        pin = [pin.strip() for pin in pincode if pin.strip().isnumeric() and len(pin.strip()) == 6]
        print(pin)
        if len(pin) != 0 :
            Report.cell(i, 6).value = pin[0]
        else:
            pincode = (pincode[len(pincode) - 1]).split('-')
            pin = [pin.strip() for pin in pincode if pin.strip().isnumeric() and len(pin.strip()) == 6]
            if len(pin) != 0 :
                Report.cell(i, 6).value = pin[0]
                
    print("Saved")
    rpt.save("Result.xlsx")
    rpt.close() 
    addrss.close()

    
main()

_____

#suffixtree

# -*- coding: utf-8 -*-
class Node:

    __num__ = -1

    def __init__(self, parentkey, outedges, suffixlink=None):
        self.parentkey = parentkey
        self.outedges = outedges
        self.suffixlink = suffixlink
        Node.__num__ += 1
        self.id = Node.__num__

    def getoutedges(self):
        return self.outedges

    def setoutedge(self, key, (anode, label_start_index, label_end_index, bnode)):
        if self.outedges is None:
            self.outedges = {}
        self.outedges[key] = (anode, label_start_index, label_end_index, bnode)

    def getoutedge(self, key):
        if key in self.outedges:
            return self.getoutedges()[key]
        else:
            return None

    def getparenkey(self):
        return self.parentkey

    def setparentkey(self, parentkey):
        self.parentkey = parentkey

    def getsuffixlink(self):
        return self.suffixlink

    def setsuffixlink(self, node):
        self.suffixlink = node

    def getid(self):
        return self.id

    @staticmethod
    def __draw__(rnode, chars, v, ed='#'):
        l = len(chars)
        edges = rnode.getoutedges().items()
        nogc = []
        hasgc = []
        gc = []
        maxlen = len(chars) + 6
        for edg in edges:
            if v == 0:
                if edg[1][3].getoutedges() is None:
                    nogc.append(edg)
                else:
                    hasgc.append(edg)
            else:
                if edg[1][3].getoutedges() is None:
                    hasgc.append(edg)
                else:
                    nogc.append(edg)
        gc.extend(hasgc)
        gc.extend(nogc)
        for k, (parent, s, t, node) in gc:
            if ed == '#':
                if t == '#':
                    t = l
            else:
                if t == '#':
                    t = ed
            linkid = ''
            if node.getsuffixlink() is not None:
                linkid = '->' + str(node.getsuffixlink().getid())

            if v == 0:
                print " " * maxlen * v + '|'
                print " " * maxlen * v + '|' + ' ' * 3 + chars[s:t + 1]
                print '+' + " " * maxlen * v + '-' + '-' * (maxlen - 1) + '● ' + '(' + str(node.getid()) + linkid + ')'
            else:
                print '|' + " " * maxlen * v + '|'
                print '|' + " " * maxlen * v + '|' + ' ' * 3 + chars[s:t + 1]
                print '|' + " " * maxlen * v + '+' + '-' * (maxlen - 1) + '● ' + '(' + str(node.getid()) + linkid + ')'
            if node.getoutedges() is not None:
                Node.__draw__(node, chars, v + 1, ed)

    @staticmethod
    def draw(root, chars, ed='#'):
        print '\n', chars, '\n● (0)'
        v = 0
        Node.__draw__(root, chars, v, ed)


def build(chars, regularize=False):
    root = Node(None, None, None)
    actnode = root
    actkey = ''
    actlen = 0
    remainder = 0  # used for splitting
    ind = 0
    while ind < len(chars):
        ch = chars[ind]
        if remainder == 0:
            if actnode.getoutedges() is not None and ch in actnode.getoutedges():
                actkey = ch
                actlen = 1
                remainder = 1
                anode, start, end, bnode = actnode.getoutedge(actkey)
                if end == '#':
                    end = ind
                if end - start + 1 == actlen:
                    actnode = actnode.getoutedge(actkey)[3]
                    actkey = ''
                    actlen = 0
            else:
                aleaf = Node(None, None, None)
                aedge = (actnode, ind, '#', aleaf)
                aleaf.setparentkey((actnode, chars[ind]))
                actnode.setoutedge(chars[ind], aedge)
        else:
            if actkey == '' and actlen == 0:  # compare on node
                if ch in actnode.getoutedges():
                    actkey = ch
                    actlen = 1
                    remainder += 1
                else:
                    remainder += 1
                    remainder, actnode, actkey, actlen = unfold(root, chars, ind, remainder, actnode, actkey, actlen)
            else:  # compare on edge
                anode, start, end, bnode = actnode.getoutedge(actkey)
                if end == '#':
                    end = ind
                compareposition = start + actlen
                if chars[compareposition] != ch:
                    remainder += 1
                    remainder, actnode, actkey, actlen = unfold(root, chars, ind, remainder, actnode, actkey, actlen)
                else:
                    if compareposition < end:  # on edge
                        actlen += 1
                        remainder += 1
                    else:  # on node
                        remainder += 1
                        actnode = actnode.getoutedge(actkey)[3]
                        if compareposition == end:
                            actlen = 0
                            actkey = ''
                        else:
                            actlen = 1
                            actkey = ch
        ind += 1
        if ind == len(chars) and remainder > 0:
            if regularize:
                chars = chars + '$'
    return root, chars


def unfold(root, chars, ind, remainder, actnode, actkey, actlen):
    prenode = None
    while remainder > 0:
        remains = chars[ind - remainder + 1:ind + 1]
        actlen_re = len(remains) - 1 - actlen
        actnode, actkey, actlen, actlen_re = hop(ind, actnode, actkey, actlen, remains, actlen_re)
        lost, actnode, actkey, actlen, actlen_re = step(chars, ind, actnode, actkey, actlen, remains, actlen_re)
        if lost:
            if actlen == 1 and prenode is not None and actnode is not root:
                prenode.setsuffixlink(actnode)
            return remainder, actnode, actkey, actlen
        if actlen == 0:
            if remains[actlen_re] not in actnode.getoutedges():
                aleaf = Node(None, None, None)
                aedge = (actnode, ind, '#', aleaf)
                aleaf.setparentkey((actnode, chars[ind]))
                actnode.setoutedge(chars[ind], aedge)
        else:  # on edge
            anode, start, end, bnode = actnode.getoutedge(actkey)
            if remains[actlen_re + actlen] != chars[start + actlen]:
                # split
                anode, start, end, bnode = actnode.getoutedge(actkey)
                newnode = Node(None, None, None)
                halfedge1 = (actnode, start, start + actlen - 1, newnode)
                halfedge2 = (newnode, start + actlen, end, bnode)
                actnode.setoutedge(actkey, halfedge1)
                newnode.setparentkey((actnode, actkey))
                newnode.setoutedge(chars[start + actlen], halfedge2)
                aleaf = Node(None, None, None)
                aedge = (newnode, ind, '#', aleaf)
                aleaf.setparentkey((newnode, chars[ind]))
                newnode.setoutedge(chars[ind], aedge)
            else:
                return remainder, actnode, actkey, actlen
        if prenode is not None and 'aleaf' in locals() and aleaf.getparenkey()[0] is not root:
            prenode.setsuffixlink(aleaf.getparenkey()[0])
        if 'aleaf' in locals() and aleaf.getparenkey()[0] is not root:
            prenode = aleaf.getparenkey()[0]
        if actnode == root and remainder > 1:
            actkey = remains[1]
            actlen -= 1
        if actnode.getsuffixlink() is not None:
            actnode = actnode.getsuffixlink()
        else:
            actnode = root
        remainder -= 1
    return remainder, actnode, actkey, actlen


def step(chars, ind, actnode, actkey, actlen, remains, ind_remainder):
    rem_label = remains[ind_remainder:]
    if actlen > 0:
        anode, start, end, bnode = actnode.getoutedge(actkey)
        if end == '#':
            end = ind
        edgelabel = chars[start:end + 1]
        if edgelabel.startswith(rem_label):
            actlen = len(rem_label)
            actkey = rem_label[0]
            return True, actnode, actkey, actlen, ind_remainder
    else:
        # on node
        if ind_remainder < len(remains) and remains[ind_remainder] in actnode.getoutedges():
            anode, start, end, bnode = actnode.getoutedge(remains[ind_remainder])
            if end == '#':
                end = ind
            edgelabel = chars[start:end + 1]
            if edgelabel.startswith(rem_label):
                actlen = len(rem_label)
                actkey = rem_label[0]
                return True, actnode, actkey, actlen, ind_remainder
    return False, actnode, actkey, actlen, ind_remainder


def hop(ind, actnode, actkey, actlen, remains, ind_remainder):
    if actlen == 0 or actkey == '':
        return actnode, actkey, actlen, ind_remainder
    anode, start, end, bnode = actnode.getoutedge(actkey)
    if end == '#':
        end = ind
    edgelength = end - start + 1
    while actlen > edgelength:
        actnode = actnode.getoutedge(actkey)[3]
        ind_remainder += edgelength
        actkey = remains[ind_remainder]
        actlen -= edgelength
        anode, start, end, bnode = actnode.getoutedge(actkey)
        if end == '#':
            end = ind
        edgelength = end - start + 1
    if actlen == edgelength:
        actnode = actnode.getoutedge(actkey)[3]
        actkey = ''
        actlen = 0
        ind_remainder += edgelength
    return actnode, actkey, actlen, ind_remainder


if __name__ == "__main__":
    docs = ['abcabxabcd', 'dedododeeodoeodooedeeododooodoede$', 'ooooooooo', 'mississippi']
    for text in docs:
        tree, pst = build(text, regularize=True)
        Node.draw(tree, pst, ed='#')
        
 _______


